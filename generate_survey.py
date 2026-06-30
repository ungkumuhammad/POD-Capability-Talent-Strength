"""
Generate the consolidated survey question master list Excel file.
Run: python3 generate_survey.py
Output: deliverables/Survey_Question_Master_List.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# DATA
# ---------------------------------------------------------------------------

QUESTIONS = [
    # ---- DEVI DUTTA --------------------------------------------------------
    {
        "contributor": "Devi Dutta",
        "theme": "Clarity & Awareness",
        "question": "How clearly do you understand your career progression path within Gentari (technical/non-technical)?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "Theme 1 of 3 from Devi",
    },
    {
        "contributor": "Devi Dutta",
        "theme": "Clarity & Awareness",
        "question": "Are you aware of any structured training or development programs available for your growth?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "",
    },
    {
        "contributor": "Devi Dutta",
        "theme": "Clarity & Awareness",
        "question": "Do you feel the expectations and criteria for promotion in your role are clearly defined?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "",
    },
    {
        "contributor": "Devi Dutta",
        "theme": "Effectiveness of Existing Programs",
        "question": "How effective are the current learning and development programs in enhancing your skills?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "Theme 2 of 3 from Devi",
    },
    {
        "contributor": "Devi Dutta",
        "theme": "Effectiveness of Existing Programs",
        "question": "Do you feel you are receiving adequate support (training, mentoring, exposure) to grow in your role?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "",
    },
    {
        "contributor": "Devi Dutta",
        "theme": "Effectiveness of Existing Programs",
        "question": "How fairly and consistently are promotion and growth opportunities applied across teams/functions?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "",
    },
    {
        "contributor": "Devi Dutta",
        "theme": "Expectations & Improvement Areas",
        "question": "What additional support or programs would help you progress in your career (e.g., mentoring, certifications, role rotations)?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "Theme 3 of 3 from Devi",
    },
    {
        "contributor": "Devi Dutta",
        "theme": "Expectations & Improvement Areas",
        "question": "Would you prefer a more clearly defined career path framework (technical vs managerial tracks)?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "",
    },
    {
        "contributor": "Devi Dutta",
        "theme": "Expectations & Improvement Areas",
        "question": "What are the key challenges you currently face in your career growth within the organization?",
        "type": "Open-ended",
        "answer_options": "",
        "audience": "Both",
        "notes": "",
    },

    # ---- FADHLI AFIQ -------------------------------------------------------
    {
        "contributor": "Fadhli Afiq",
        "theme": "Validate Current State",
        "question": "How clear is your career development pathway within Gentari for your current role (technical or non-technical)?",
        "type": "Rating Scale (Likert)",
        "answer_options": "Very clear | Somewhat clear | Neutral | Somewhat unclear | Very unclear",
        "audience": "Both",
        "notes": "Q1 of 3 from Fadhli",
    },
    {
        "contributor": "Fadhli Afiq",
        "theme": "Validate Customer Pain",
        "question": "Which of the following best describes the biggest challenge you face in developing your career at Gentari? (Select up to two)",
        "type": "MCQ (Multi-select)",
        "answer_options": (
            "I do not know what competencies are required for the next role | "
            "I do not know what training or development I should pursue | "
            "There is no structured career pathway for my role | "
            "Opportunities for technical growth are limited | "
            "Opportunities for leadership/non-technical growth are limited | "
            "Promotion criteria are unclear | "
            "I do not face significant challenges | "
            "Other"
        ),
        "audience": "Both",
        "notes": "Q2 of 3 from Fadhli",
    },
    {
        "contributor": "Fadhli Afiq",
        "theme": "Validate Desired Outcome",
        "question": "If Gentari could improve one aspect of career development, which would create the greatest value for you?",
        "type": "MCQ (Single-select)",
        "answer_options": (
            "A clear career progression roadmap for my role | "
            "A competency framework showing skills required for each career level | "
            "Structured learning and certification pathways | "
            "Individual development plans with regular career discussions | "
            "Mentoring or coaching from experienced colleagues | "
            "Greater visibility of internal career opportunities | "
            "Other"
        ),
        "audience": "Both",
        "notes": "Q3 of 3 from Fadhli",
    },
]

# ---------------------------------------------------------------------------
# STYLE HELPERS
# ---------------------------------------------------------------------------

TEAL       = "006B6B"   # header background
WHITE      = "FFFFFF"
LIGHT_TEAL = "E6F4F4"   # alternating row
DARK_TEAL  = "004F4F"   # header font

CONTRIBUTOR_COLORS = {
    "Devi Dutta":  "D9EAD3",   # soft green
    "Fadhli Afiq": "CFE2F3",   # soft blue
}

THIN = Side(border_style="thin", color="B0C4C4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    "No.",
    "Contributor",
    "Theme",
    "Question",
    "Type",
    "Answer Options",
    "Audience",
    "Notes",
]

COL_WIDTHS = [6, 18, 30, 70, 24, 80, 14, 30]


def header_style(cell):
    cell.font      = Font(bold=True, color=WHITE, size=11)
    cell.fill      = PatternFill("solid", fgColor=TEAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def data_style(cell, contributor, col_idx, row_shade):
    bg = CONTRIBUTOR_COLORS.get(contributor, WHITE) if not row_shade else LIGHT_TEAL
    # contributor colour takes priority over generic shading
    bg = CONTRIBUTOR_COLORS.get(contributor, LIGHT_TEAL if row_shade else WHITE)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border    = BORDER
    if col_idx == 1:          # No. column centred
        cell.alignment = Alignment(horizontal="center", vertical="top")


# ---------------------------------------------------------------------------
# BUILD WORKBOOK
# ---------------------------------------------------------------------------

def build_workbook(output_path: str):
    wb = Workbook()

    # ---- SHEET 1: Master Question List ------------------------------------
    ws = wb.active
    ws.title = "Master Question List"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        header_style(cell)

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    current_contributor = None
    color_toggle = False

    for row_idx, q in enumerate(QUESTIONS, start=2):
        if q["contributor"] != current_contributor:
            current_contributor = q["contributor"]
            color_toggle = not color_toggle

        row_data = [
            row_idx - 1,
            q["contributor"],
            q["theme"],
            q["question"],
            q["type"],
            q["answer_options"],
            q["audience"],
            q["notes"],
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            data_style(cell, q["contributor"], col_idx, color_toggle)

        ws.row_dimensions[row_idx].height = 60

    # ---- SHEET 2: Summary ------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16

    summary_data = [
        ("Metric", "Count"),
        ("Total questions", len(QUESTIONS)),
        ("Contributors", len({q["contributor"] for q in QUESTIONS})),
        ("Devi Dutta – questions", sum(1 for q in QUESTIONS if q["contributor"] == "Devi Dutta")),
        ("Fadhli Afiq – questions", sum(1 for q in QUESTIONS if q["contributor"] == "Fadhli Afiq")),
        ("Open-ended", sum(1 for q in QUESTIONS if q["type"] == "Open-ended")),
        ("Rating Scale (Likert)", sum(1 for q in QUESTIONS if "Rating" in q["type"])),
        ("MCQ (Multi-select)", sum(1 for q in QUESTIONS if "Multi" in q["type"])),
        ("MCQ (Single-select)", sum(1 for q in QUESTIONS if "Single" in q["type"])),
        ("Audience: Both", sum(1 for q in QUESTIONS if q["audience"] == "Both")),
    ]

    for r_idx, (label, value) in enumerate(summary_data, start=1):
        c_label = ws2.cell(row=r_idx, column=1, value=label)
        c_value = ws2.cell(row=r_idx, column=2, value=value)
        if r_idx == 1:
            header_style(c_label)
            header_style(c_value)
        else:
            bg = LIGHT_TEAL if r_idx % 2 == 0 else WHITE
            for cell in (c_label, c_value):
                cell.fill   = PatternFill("solid", fgColor=bg)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center")
            c_value.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r_idx].height = 22

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    build_workbook("deliverables/Survey_Question_Master_List.xlsx")
